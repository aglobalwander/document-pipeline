"""Utility for creating Excel templates."""
import os
import logging
import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional, List

class ExcelTemplateCreator:
    """Create Excel templates for use with JsonToExcel transformer."""

    def __init__(self, template_dir: str = 'report_templates/excel'):
        """Initialize the Excel template creator.
        
        Args:
            template_dir: Directory to store templates
        """
        self.logger = logging.getLogger(__name__)
        self.template_dir = template_dir
        
        # Ensure template directory exists
        os.makedirs(self.template_dir, exist_ok=True)
        
        self.logger.info(f"Initialized ExcelTemplateCreator (template_dir={self.template_dir})")

    def create_default_template(self) -> str:
        """Create a default Excel template with a blank sheet and header row formatting.
        
        Returns:
            Path to the created template
        """
        template_path = Path(self.template_dir) / "default.xlsx"
        
        try:
            # Create a workbook with a single sheet
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Data"
            
            # Add a header row with bold formatting
            headers = ["Column1", "Column2", "Column3"]
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(header) + 2
                
            # Save the template
            wb.save(template_path)
            self.logger.info(f"Created default template at {template_path}")
            
            return str(template_path)
            
        except Exception as e:
            self.logger.error(f"Error creating default template: {e}")
            return ""

    def create_custom_template(self, template_name: str, sheets: List[Dict[str, Any]]) -> str:
        """Create a custom Excel template with multiple sheets and formatting.
        
        Args:
            template_name: Name of the template (without extension)
            sheets: List of sheet configurations, each with:
                - name: Sheet name
                - headers: List of column headers
                - column_widths: Optional dict of column indices to widths
                - freeze_panes: Optional cell reference to freeze panes (e.g., 'B2')
                - styles: Optional dict of style configurations
                
        Returns:
            Path to the created template
        """
        template_path = Path(self.template_dir) / f"{template_name}.xlsx"
        
        try:
            # Create a workbook
            wb = openpyxl.Workbook()
            
            # Remove the default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Create each sheet
            for sheet_config in sheets:
                sheet_name = sheet_config.get('name', 'Sheet')
                headers = sheet_config.get('headers', [])
                column_widths = sheet_config.get('column_widths', {})
                freeze_panes = sheet_config.get('freeze_panes')
                styles = sheet_config.get('styles', {})
                
                # Create the sheet
                sheet = wb.create_sheet(sheet_name)
                
                # Add headers
                for col_num, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = openpyxl.styles.Font(bold=True)
                    
                # Set column widths
                for col_num, width in column_widths.items():
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(int(col_num))].width = width
                    
                # Set freeze panes
                if freeze_panes:
                    sheet.freeze_panes = sheet[freeze_panes]
                    
                # Apply styles
                self._apply_styles(sheet, styles)
                
            # Save the template
            wb.save(template_path)
            self.logger.info(f"Created custom template '{template_name}' at {template_path}")
            
            return str(template_path)
            
        except Exception as e:
            self.logger.error(f"Error creating custom template '{template_name}': {e}")
            return ""

    def _apply_styles(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                     styles: Dict[str, Any]) -> None:
        """Apply styles to a worksheet.
        
        Args:
            sheet: Worksheet to apply styles to
            styles: Style configurations
        """
        # Apply table style
        if 'table_style' in styles:
            # Note: Full table styling requires more complex code with openpyxl
            # This is a simplified version
            table_style = styles['table_style']
            if table_style.get('alternating_rows', False):
                # Apply alternating row colors
                for row_idx in range(2, 100):  # Pre-style 100 rows
                    if row_idx % 2 == 0:
                        for col_idx in range(1, len(sheet[1]) + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="F2F2F2", 
                                end_color="F2F2F2",
                                fill_type="solid"
                            )
        
        # Apply header style
        if 'header_style' in styles:
            header_style = styles['header_style']
            header_bg_color = header_style.get('background_color', '4472C4')
            header_font_color = header_style.get('font_color', 'FFFFFF')
            
            for cell in sheet[1]:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color=header_bg_color, 
                    end_color=header_bg_color,
                    fill_type="solid"
                )
                cell.font = openpyxl.styles.Font(
                    bold=True, 
                    color=header_font_color
                )
                
        # Apply cell styles
        if 'cell_styles' in styles:
            for cell_ref, cell_style in styles['cell_styles'].items():
                try:
                    cell = sheet[cell_ref]
                    
                    if 'font' in cell_style:
                        font_style = cell_style['font']
                        cell.font = openpyxl.styles.Font(
                            bold=font_style.get('bold', False),
                            italic=font_style.get('italic', False),
                            color=font_style.get('color')
                        )
                        
                    if 'fill' in cell_style:
                        fill_style = cell_style['fill']
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color=fill_style.get('color', 'FFFFFF'),
                            end_color=fill_style.get('color', 'FFFFFF'),
                            fill_type="solid"
                        )
                        
                    if 'alignment' in cell_style:
                        alignment_style = cell_style['alignment']
                        cell.alignment = openpyxl.styles.Alignment(
                            horizontal=alignment_style.get('horizontal'),
                            vertical=alignment_style.get('vertical')
                        )
                except Exception as e:
                    self.logger.error(f"Error applying cell style to {cell_ref}: {e}")