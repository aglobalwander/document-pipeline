"""Create default Excel template for the pipeline."""
import os
import openpyxl
from pathlib import Path

def create_default_template():
    """Create a default Excel template with a blank sheet and header row formatting."""
    template_dir = 'report_templates/excel'
    template_path = Path(template_dir) / "default.xlsx"
    
    # Ensure template directory exists
    os.makedirs(template_dir, exist_ok=True)
    
    try:
        # Create a workbook with a single sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Data"
        
        # Add a header row with bold formatting
        headers = ["Column1", "Column2", "Column3"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(header) + 2
            
        # Save the template
        wb.save(template_path)
        print(f"Created default template at {template_path}")
        
        return str(template_path)
        
    except Exception as e:
        print(f"Error creating default template: {e}")
        return ""

if __name__ == "__main__":
    create_default_template()