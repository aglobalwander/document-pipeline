"""Transform JSON data to Excel format."""
import os
import json
import logging
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional, List, Union

from doc_processing.embedding.base import PipelineComponent
from doc_processing.embedding.base import BaseTransformer
from doc_processing.config import get_settings

class JsonToExcel(BaseTransformer, PipelineComponent):
    """Transform JSON data to Excel format."""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize the JSON to Excel transformer.
        
        Args:
            config: Configuration options
                excel_template: Name of Excel template (default: 'default')
                excel_template_dir: Directory containing Excel templates (default: 'report_templates/excel')
        """
        super().__init__(config or {})
        self.logger = logging.getLogger(__name__)
        self.settings = get_settings()
        
        # Get template configuration
        self.template_name = self.config.get('excel_template', 'default')
        self.template_dir = self.config.get('excel_template_dir', 'report_templates/excel')
        
        # Ensure template directory exists
        os.makedirs(self.template_dir, exist_ok=True)
        
        self.logger.info(f"Initialized JsonToExcel transformer (template={self.template_name})")

    def transform(self, document: Dict[str, Any]) -> Dict[str, Any]:
        """Transform JSON data to Excel.
        
        Args:
            document: Document dictionary with content field containing JSON data
            
        Returns:
            Processed document with Excel output paths
        """
        if not document.get('content'):
            self.logger.error("No content field in document")
            return document
            
        # Get the input path from metadata if available
        input_path = document.get('metadata', {}).get('input_path', '')
        if not input_path:
            self.logger.warning("No input_path in metadata, using generic filename")
            input_path = "data.json"
            
        # Create output directory
        output_dir = Path(self.settings.OUTPUT_DIR) / "xlsx"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate output filename
        input_filename = Path(input_path).stem
        output_path = output_dir / f"{input_filename}.xlsx"
        
        try:
            # Parse JSON content
            if isinstance(document['content'], str):
                try:
                    json_data = json.loads(document['content'])
                except json.JSONDecodeError:
                    self.logger.error("Failed to parse JSON content")
                    return document
            else:
                json_data = document['content']
                
            # Convert JSON to DataFrame
            df = self._json_to_dataframe(json_data)
            
            # Save to Excel
            excel_path = self._save_to_excel(df, output_path, input_filename)
            
            # Update document
            result = document.copy()
            result['excel_path'] = excel_path
            result['metadata'] = document.get('metadata', {})
            result['metadata']['excel_output'] = excel_path
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error converting JSON to Excel: {e}")
            return document

    def _json_to_dataframe(self, json_data: Any) -> Dict[str, pd.DataFrame]:
        """Convert JSON data to pandas DataFrames.
        
        Args:
            json_data: JSON data (dict, list, or other JSON-serializable object)
            
        Returns:
            Dictionary of sheet names to pandas DataFrames
        """
        dataframes = {}
        
        # Handle different JSON structures
        if isinstance(json_data, list):
            # List of objects/records
            dataframes['Data'] = pd.json_normalize(json_data)
        elif isinstance(json_data, dict):
            # Check if it's a multi-sheet structure
            if all(isinstance(value, (list, dict)) for value in json_data.values()):
                # Each top-level key becomes a sheet
                for key, value in json_data.items():
                    sheet_name = self._sanitize_sheet_name(key)
                    if isinstance(value, list):
                        dataframes[sheet_name] = pd.json_normalize(value)
                    elif isinstance(value, dict):
                        # Flatten one level of nesting
                        flattened = {}
                        for subkey, subvalue in value.items():
                            if isinstance(subvalue, dict):
                                for subsubkey, subsubvalue in subvalue.items():
                                    flattened[f"{subkey}.{subsubkey}"] = subsubvalue
                            else:
                                flattened[subkey] = subvalue
                        dataframes[sheet_name] = pd.DataFrame([flattened])
            else:
                # Single object or nested structure
                # Flatten one level of nesting
                flattened = {}
                for key, value in json_data.items():
                    if isinstance(value, dict):
                        for subkey, subvalue in value.items():
                            flattened[f"{key}.{subkey}"] = subvalue
                    else:
                        flattened[key] = value
                dataframes['Data'] = pd.DataFrame([flattened])
        else:
            # Scalar value or other type
            self.logger.warning(f"Unexpected JSON structure: {type(json_data)}")
            dataframes['Data'] = pd.DataFrame([{"value": json_data}])
            
        return dataframes

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name to be valid for Excel.
        
        Args:
            name: Sheet name
            
        Returns:
            Sanitized sheet name
        """
        # Excel sheet names cannot exceed 31 characters
        name = str(name)[:31]
        
        # Replace invalid characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
            
        return name

    def _save_to_excel(self, dataframes: Dict[str, pd.DataFrame], output_path: Path,
                       input_filename: str) -> str:
        """Save DataFrames to Excel file using template.

        Args:
            dataframes: Dictionary of sheet names to pandas DataFrames
            output_path: Path to save the Excel file
            input_filename: Original input filename (without extension)

        Returns:
            Path to the Excel file
        """
        template_path = Path(self.template_dir) / f"{self.template_name}.xlsx"

        try:
            # Load the template workbook if it exists, otherwise create a new one
            if template_path.exists():
                self.logger.info(f"Loading template from {template_path}")
                output_wb = openpyxl.load_workbook(template_path)
            else:
                self.logger.warning(f"Template {template_path} not found, creating a new workbook.")
                output_wb = openpyxl.Workbook()
                # Remove the default sheet created by openpyxl.Workbook()
                if 'Sheet' in output_wb.sheetnames:
                    output_wb.remove(output_wb['Sheet'])

            # Use a Pandas ExcelWriter with the loaded workbook
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                writer.book = output_wb

                # Write each DataFrame to a sheet
                for sheet_name, df in dataframes.items():
                    # Skip empty DataFrames
                    if df.empty:
                        continue

                    # Create a new sheet if it doesn't exist in the template
                    if sheet_name not in writer.book.sheetnames:
                        writer.book.create_sheet(sheet_name)

                    # Write DataFrame to sheet, starting from the second row if headers are present in template
                    # This simple approach overwrites existing content in the sheet.
                    # For more complex template usage (e.g., writing to specific cells),
                    # direct openpyxl manipulation would be needed.
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Apply formatting (using the existing method, which now operates on the writer's book)
                    self._apply_formatting(writer, sheet_name, df)

                # Save the Excel file
                writer.close()

            self.logger.info(f"Saved Excel file to {output_path}")
            return str(output_path)

        except Exception as e:
            self.logger.error(f"Error saving Excel file with template: {e}")
            # Fallback to simple Excel export without template
            try:
                self.logger.warning("Attempting fallback Excel export without template.")
                with pd.ExcelWriter(output_path) as writer:
                    for sheet_name, df in dataframes.items():
                        if not df.empty:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                self.logger.info(f"Saved Excel file (fallback) to {output_path}")
                return str(output_path)
            except Exception as e2:
                self.logger.error(f"Error in fallback Excel export: {e2}")
                return ""

    def _create_default_template(self, template_path: Path) -> None:
        """Create a default Excel template.

        Args:
            template_path: Path to save the template
        """
        try:
            # Create a workbook with a single sheet
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Data"

            # Add a header row with bold formatting (basic example)
            headers = ["Column1", "Column2", "Column3"] # Example headers
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)

            # Save the template
            wb.save(template_path)
            self.logger.info(f"Created default template at {template_path}")
        except Exception as e:
            self.logger.error(f"Error creating default template: {e}")

    def _apply_formatting(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
        """Apply formatting to Excel sheet.

        Args:
            writer: Excel writer
            sheet_name: Sheet name
            df: DataFrame
        """
        try:
            # Get the worksheet from the writer's book
            worksheet = writer.book[sheet_name]

            # Format header row (bold) - assuming headers are in the first row
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = openpyxl.styles.Font(bold=True)

            # Auto-adjust column widths
            for col_num, column in enumerate(df.columns, 1):
                max_length = 0
                column_name = str(column)
                max_length = max(max_length, len(column_name))

                # Check data in the column
                # Iterate through rows starting from the second row (after header)
                for row_index in range(len(df)):
                     cell_value = df.iloc[row_index, col_num - 1] # Use iloc for row/column index
                     max_length = max(max_length, len(str(cell_value) if cell_value is not None else ''))


                # Adjust column width (with a buffer)
                adjusted_width = max_length + 2
                # Ensure width is not excessively large
                adjusted_width = min(adjusted_width, 50) # Cap width at 50 for readability
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width

        except Exception as e:
            self.logger.error(f"Error applying formatting: {e}")